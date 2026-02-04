import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
import openpyxl
import os
import sys
import platform # Import platform untuk deteksi OS
import subprocess # Import subprocess untuk membuka file di non-Windows (jaga-jaga)
from datetime import datetime
from PIL import Image, ImageTk

# --- FUNGSI SAKTI: RESOURCE PATH ---
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# --- KONFIGURASI FILE ---
NAMA_FILE = "pembukuan_up_kpu_tubaba.xlsx"
PAGU_TETAP = 11400000
LOGO_FILENAME = resource_path("kpu_logo.png") 

# --- COLOR PALETTE ---
KPU_RED_DARK = "#B71C1C"    
KPU_RED_LIGHT = "#D32F2F"   
BG_MAIN = "#F5F5F5"         
BG_WHITE = "#FFFFFF"
TEXT_DARK = "#212121"       
TEXT_WHITE = "#FFFFFF"
SUCCESS_GREEN = "#2E7D32" # Hijau Excel
TABLE_STRIPE_BG = "#F0F0F0" 
TABLE_SELECTED = "#E0E0E0"  

id_yang_diedit = None 
photo_logo = None 

def format_rupiah(angka):
    return f"Rp {angka:,.0f}".replace(",", ".")

def inisialisasi_excel():
    if not os.path.exists(NAMA_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data Pengeluaran"
        headers = ["No", "Tanggal", "Jenis", "Uraian", "Nominal", "No. Rekening", "Penerima", "Penyedia", "Kode Akun"]
        ws.append(headers)
        dims = {'A':5, 'B':12, 'C':10, 'D':35, 'E':18, 'F':20, 'G':20, 'H':20, 'I':20}
        for col, width in dims.items():
            ws.column_dimensions[col].width = width
        wb.save(NAMA_FILE)

def hitung_posisi_keuangan():
    total_belanja_all = 0
    total_gup_all = 0
    total_belanja_bulan_ini = 0
    belanja_pending_gup = 0 
    
    bulan_ini = datetime.now().month
    tahun_ini = datetime.now().year

    if os.path.exists(NAMA_FILE):
        wb = openpyxl.load_workbook(NAMA_FILE)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        
        for row in rows:
            try:
                tgl_str, jenis, nominal = row[1], row[2], row[4]
                if isinstance(nominal, (int, float)):
                    if jenis == "BELANJA":
                        total_belanja_all += nominal
                        try:
                            tgl_obj = datetime.strptime(tgl_str, "%d-%m-%Y")
                            if tgl_obj.month == bulan_ini and tgl_obj.year == tahun_ini:
                                total_belanja_bulan_ini += nominal
                        except: pass
                    elif jenis == "GUP":
                        total_gup_all += nominal
            except: pass

        for row in reversed(rows):
            try:
                jenis, nominal = row[2], row[4]
                if jenis == "GUP": break
                elif jenis == "BELANJA" and isinstance(nominal, (int, float)):
                    belanja_pending_gup += nominal
            except: pass

    saldo_fisik = PAGU_TETAP - total_belanja_all + total_gup_all
    return saldo_fisik, total_belanja_bulan_ini, belanja_pending_gup

def proses_gup_otomatis():
    saldo, bulan_ini, gup_needed = hitung_posisi_keuangan()
    if gup_needed <= 0:
        messagebox.showinfo("Info GUP", "Belum ada pengeluaran baru yang perlu diganti (Revolving).\nSaldo masih utuh sesuai GUP terakhir.")
        return
    pesan = (f"Sistem mendeteksi pemakaian dana sebesar {format_rupiah(gup_needed)}.\n\n"
             f"Apakah dana GUP sejumlah tersebut sudah masuk ke rekening/kas?\n"
             f"Klik YES untuk mereset saldo kembali ke Pagu Rp {format_rupiah(PAGU_TETAP)}.")
    
    if messagebox.askyesno("Konfirmasi Terima GUP", pesan):
        try:
            wb = openpyxl.load_workbook(NAMA_FILE)
            ws = wb.active
            no_urut = ws.max_row
            tgl = datetime.now().strftime("%d-%m-%Y")
            ws.append([no_urut, tgl, "GUP", "Terima GUP (Revolving Fund)", gup_needed, "-", "Bendahara", "-", "-"])
            wb.save(NAMA_FILE)
            messagebox.showinfo("Sukses", f"Saldo berhasil diisi kembali!\nSaldo Kas sekarang: Rp {format_rupiah(PAGU_TETAP)}")
            refresh_tampilan()
        except PermissionError:
             messagebox.showerror("Gagal", "Tutup file Excel dulu!")

# --- FUNGSI BARU: BUKA EXCEL ---
def buka_file_excel():
    if not os.path.exists(NAMA_FILE):
        messagebox.showerror("Error", "File database belum ditemukan!")
        return
    
    try:
        if platform.system() == 'Windows':
            os.startfile(NAMA_FILE)
        elif platform.system() == 'Darwin':       # macOS
            subprocess.call(('open', NAMA_FILE))
        else:                                   # Linux
            subprocess.call(('xdg-open', NAMA_FILE))
    except Exception as e:
        messagebox.showerror("Error", f"Gagal membuka file Excel: {e}")

def simpan_transaksi():
    simpan_data("BELANJA")

def simpan_data(jenis_transaksi):
    global id_yang_diedit
    tgl = entry_tanggal.get()
    uraian = entry_uraian.get("1.0", tk.END).strip()
    nominal_str = entry_pengeluaran.get()
    rek = entry_rek.get()
    penerima = entry_penerima.get()
    penyedia = entry_penyedia.get()
    akun = entry_akun.get()

    if not tgl or not uraian or not nominal_str:
        messagebox.showwarning("Peringatan", "Tanggal, Uraian, dan Nominal wajib diisi!")
        return

    try:
        nominal = int(nominal_str)
    except ValueError:
        messagebox.showerror("Error", "Nominal harus angka (tanpa titik/koma)!")
        return

    try:
        wb = openpyxl.load_workbook(NAMA_FILE)
        ws = wb.active

        if id_yang_diedit is None:
            no_urut = ws.max_row 
            ws.append([no_urut, tgl, jenis_transaksi, uraian, nominal, rek, penerima, penyedia, akun])
            pesan = "Transaksi berhasil direkam!"
        else:
            found = False
            for row in ws.iter_rows(min_row=2):
                if str(row[0].value) == str(id_yang_diedit):
                    row[1].value = tgl
                    row[3].value = uraian
                    row[4].value = nominal
                    row[5].value = rek
                    row[6].value = penerima
                    row[7].value = penyedia
                    row[8].value = akun
                    found = True
                    break
            if not found:
                messagebox.showerror("Error", "Data tidak ditemukan."); return
            pesan = "Data berhasil diupdate!"
            id_yang_diedit = None
            btn_simpan.config(text="SIMPAN PENGELUARAN", bg=KPU_RED_DARK)
            btn_batal.grid_remove()

        wb.save(NAMA_FILE)
        messagebox.showinfo("Sukses", pesan)
        bersihkan_form()
        refresh_tampilan()
    except PermissionError:
        messagebox.showerror("Gagal", "Tutup file Excel dulu!")

def mode_edit():
    global id_yang_diedit
    selected = tabel.selection()
    if not selected: return
    val = tabel.item(selected, 'values')
    if val[2] == "GUP":
        messagebox.showinfo("Akses Ditolak", "Data GUP adalah data sistem. Tidak bisa diedit manual.")
        return
    id_yang_diedit = val[0]
    entry_tanggal.delete(0, tk.END); entry_tanggal.insert(0, val[1])
    entry_uraian.delete("1.0", tk.END); entry_uraian.insert("1.0", val[3])
    nom = val[4].replace("Rp ", "").replace(".", "")
    entry_pengeluaran.delete(0, tk.END); entry_pengeluaran.insert(0, nom)
    entry_rek.delete(0, tk.END); entry_rek.insert(0, val[5])
    entry_penerima.delete(0, tk.END); entry_penerima.insert(0, val[6])
    entry_penyedia.delete(0, tk.END); entry_penyedia.insert(0, val[7])
    entry_akun.delete(0, tk.END); entry_akun.insert(0, val[8])
    btn_simpan.config(text="UPDATE DATA", bg="#FF9800")
    btn_batal.grid(row=4, column=4, padx=5)

def batalkan_edit():
    global id_yang_diedit
    id_yang_diedit = None
    bersihkan_form()
    btn_simpan.config(text="SIMPAN PENGELUARAN", bg=KPU_RED_DARK)
    btn_batal.grid_remove()

def bersihkan_form():
    entry_uraian.delete("1.0", tk.END)
    entry_pengeluaran.delete(0, tk.END)
    entry_rek.delete(0, tk.END)
    entry_penerima.delete(0, tk.END)
    entry_penyedia.delete(0, tk.END)

def refresh_tampilan():
    saldo, pemakaian_bulan_ini, gup_pending = hitung_posisi_keuangan()
    lbl_saldo.config(text=f"{format_rupiah(saldo)}")
    lbl_pemakaian.config(text=f"{format_rupiah(pemakaian_bulan_ini)}")
    
    if saldo < (PAGU_TETAP * 0.2):
        lbl_saldo_title.config(fg=KPU_RED_DARK)
        lbl_saldo.config(fg=KPU_RED_DARK)
    else:
        lbl_saldo_title.config(fg=SUCCESS_GREEN)
        lbl_saldo.config(fg=SUCCESS_GREEN)

    if gup_pending >= (PAGU_TETAP * 0.5):
        btn_gup.config(bg="#FF9800", text=f"⚠ GUP READY: {format_rupiah(gup_pending)}")
    else:
        btn_gup.config(bg=KPU_RED_DARK, text=f"+ TERIMA GUP (Pending: {format_rupiah(gup_pending)})")

    for item in tabel.get_children():
        tabel.delete(item)
        
    if os.path.exists(NAMA_FILE):
        wb = openpyxl.load_workbook(NAMA_FILE)
        ws = wb.active
        data = list(ws.iter_rows(min_row=2, values_only=True))
        for i, row in enumerate(reversed(data)):
            r = list(row)
            if isinstance(r[4], (int, float)):
                r[4] = format_rupiah(r[4])
            tag_warna = 'evenrow' if i % 2 == 0 else 'oddrow'
            if r[2] == "GUP": tag_warna = 'guprow'
            tabel.insert("", tk.END, values=r, tags=(tag_warna,))

# --- GUI SETUP ---
inisialisasi_excel()
root = tk.Tk()
root.title("Aplikasi Monitoring GUP - Satker KPU Tulang Bawang Barat")
root.state('zoomed')
root.configure(bg=BG_MAIN)

try:
    icon_img = tk.PhotoImage(file=LOGO_FILENAME)
    root.iconphoto(False, icon_img)
except Exception: pass

style = ttk.Style()
style.theme_use("clam")
style.configure("Treeview.Heading", background=KPU_RED_DARK, foreground=TEXT_WHITE, font=("Arial", 10, "bold"), relief="flat")
style.configure("Treeview", background=BG_WHITE, fieldbackground=BG_WHITE, foreground=TEXT_DARK, rowheight=28, font=("Arial", 10))
style.map('Treeview', background=[('selected', TABLE_SELECTED)], foreground=[('selected', 'black')])

frame_top_banner = tk.Frame(root, bg=KPU_RED_DARK, pady=10)
frame_top_banner.pack(fill=tk.X)
header_content = tk.Frame(frame_top_banner, bg=KPU_RED_DARK)
header_content.pack()

try:
    img = Image.open(LOGO_FILENAME)
    img = img.resize((60, 60), Image.LANCZOS) 
    photo_logo = ImageTk.PhotoImage(img)
    logo_label = tk.Label(header_content, image=photo_logo, bg=KPU_RED_DARK)
    logo_label.pack(side=tk.LEFT, padx=(0, 15))
except FileNotFoundError:
    tk.Label(header_content, text="[LOGO KPU]", bg=KPU_RED_DARK, fg=TEXT_WHITE).pack(side=tk.LEFT, padx=(0, 15))

tk.Label(header_content, text="KOMISI PEMILIHAN UMUM", bg=KPU_RED_DARK, fg=TEXT_WHITE, font=("Arial", 14)).pack(anchor="w")
tk.Label(header_content, text="KABUPATEN TULANG BAWANG BARAT", bg=KPU_RED_DARK, fg=TEXT_WHITE, font=("Arial", 18, "bold")).pack(anchor="w")

frame_info = tk.Frame(root, bg=BG_WHITE, bd=1, relief=tk.SOLID)
frame_info.pack(fill=tk.X, padx=15, pady=15)
frame_saldo = tk.Frame(frame_info, bg=BG_WHITE, padx=20, pady=10)
frame_saldo.pack(side=tk.LEFT)
lbl_saldo_title = tk.Label(frame_saldo, text="SALDO KAS (REAL)", bg=BG_WHITE, font=("Arial", 10, "bold"), fg=SUCCESS_GREEN)
lbl_saldo_title.pack(anchor="w")
lbl_saldo = tk.Label(frame_saldo, text="Rp 0", bg=BG_WHITE, font=("Arial", 22, "bold"), fg=SUCCESS_GREEN)
lbl_saldo.pack(anchor="w")
ttk.Separator(frame_info, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=10, pady=5)
frame_pakai = tk.Frame(frame_info, bg=BG_WHITE, padx=20, pady=10)
frame_pakai.pack(side=tk.LEFT)
tk.Label(frame_pakai, text="REALISASI BELANJA (BULAN INI)", bg=BG_WHITE, font=("Arial", 10, "bold"), fg=KPU_RED_LIGHT).pack(anchor="w")
lbl_pemakaian = tk.Label(frame_pakai, text="Rp 0", bg=BG_WHITE, font=("Arial", 22, "bold"), fg=KPU_RED_LIGHT)
lbl_pemakaian.pack(anchor="w")
btn_gup = tk.Button(frame_info, text="Checking...", command=proses_gup_otomatis, bg=KPU_RED_DARK, fg=TEXT_WHITE, font=("Arial", 11, "bold"), height=2, padx=25, relief=tk.FLAT, cursor="hand2")
btn_gup.pack(side=tk.RIGHT, padx=20, pady=15)

frame_input = tk.LabelFrame(root, text=" Input Transaksi Belanja ", padx=15, pady=15, bg=BG_MAIN, font=("Arial", 11, "bold"), fg=KPU_RED_DARK, bd=2, relief=tk.GROOVE)
frame_input.pack(fill=tk.X, padx=15)

def style_label(text): return tk.Label(frame_input, text=text, bg=BG_MAIN, font=("Arial", 10))
def style_entry(width): return tk.Entry(frame_input, width=width, font=("Arial", 10), bd=2, relief=tk.FLAT)

style_label("Tanggal:").grid(row=0, column=0, sticky="w", pady=5)
entry_tanggal = style_entry(20)
entry_tanggal.insert(0, datetime.now().strftime("%d-%m-%Y"))
entry_tanggal.grid(row=0, column=1, padx=10, pady=5)
style_label("Nominal (Rp):").grid(row=0, column=2, sticky="w", pady=5)
entry_pengeluaran = style_entry(20)
entry_pengeluaran.grid(row=0, column=3, padx=10, pady=5)
style_label("Uraian:").grid(row=1, column=0, sticky="nw", pady=5)
entry_uraian = tk.Text(frame_input, height=3, width=55, font=("Arial", 10), bd=2, relief=tk.FLAT)
entry_uraian.grid(row=1, column=1, columnspan=3, sticky="we", padx=10, pady=5)
style_label("No. Rekening:").grid(row=2, column=0, sticky="w", pady=5)
entry_rek = style_entry(20)
entry_rek.grid(row=2, column=1, padx=10, pady=5)
style_label("Penerima:").grid(row=2, column=2, sticky="w", pady=5)
entry_penerima = style_entry(20)
entry_penerima.grid(row=2, column=3, padx=10, pady=5)
style_label("Penyedia:").grid(row=3, column=0, sticky="w", pady=5)
entry_penyedia = style_entry(20)
entry_penyedia.grid(row=3, column=1, padx=10, pady=5)
style_label("Kode Akun:").grid(row=3, column=2, sticky="w", pady=5)
entry_akun = style_entry(35)
entry_akun.grid(row=3, column=3, padx=10, pady=5)

btn_simpan = tk.Button(frame_input, text="SIMPAN PENGELUARAN", command=simpan_transaksi, bg=KPU_RED_DARK, fg=TEXT_WHITE, font=("Arial", 10, "bold"), padx=15, pady=5, relief=tk.FLAT, cursor="hand2")
btn_simpan.grid(row=4, column=3, sticky="e", padx=10, pady=15)
btn_batal = tk.Button(frame_input, text="BATAL EDIT", command=batalkan_edit, bg=TEXT_DARK, fg=TEXT_WHITE, font=("Arial", 10, "bold"), padx=15, pady=5, relief=tk.FLAT, cursor="hand2")

frame_tabel = tk.Frame(root, bg=BG_MAIN)
frame_tabel.pack(fill=tk.BOTH, expand=True, padx=15, pady=(0, 15))

# --- AREA TOMBOL DI ATAS TABEL ---
frame_tombol_tabel = tk.Frame(frame_tabel, bg=BG_MAIN)
frame_tombol_tabel.pack(fill=tk.X, pady=(0, 5))

# Tombol EDIT (Kiri)
btn_edit = tk.Button(frame_tombol_tabel, text="EDIT DATA TERPILIH", command=mode_edit, bg=KPU_RED_LIGHT, fg=TEXT_WHITE, font=("Arial", 9, "bold"), relief=tk.FLAT, padx=10, cursor="hand2")
btn_edit.pack(side=tk.LEFT)

# Tombol BUKA EXCEL (Kanan)
btn_excel = tk.Button(frame_tombol_tabel, text="📂 BUKA FILE EXCEL", command=buka_file_excel, bg=SUCCESS_GREEN, fg=TEXT_WHITE, font=("Arial", 9, "bold"), relief=tk.FLAT, padx=10, cursor="hand2")
btn_excel.pack(side=tk.RIGHT)

cols = ("No", "Tanggal", "Jenis", "Uraian", "Nominal", "Rek", "Penerima", "Penyedia", "Akun")
tabel = ttk.Treeview(frame_tabel, columns=cols, show="headings", height=15)
tabel.tag_configure('oddrow', background=BG_WHITE)
tabel.tag_configure('evenrow', background=TABLE_STRIPE_BG) 
tabel.tag_configure('guprow', background="#E1BEE7", font=("Arial", 10, "bold")) 

tabel.heading("No", text="No"); tabel.column("No", width=50, anchor=tk.CENTER)
tabel.heading("Tanggal", text="Tanggal"); tabel.column("Tanggal", width=90, anchor=tk.CENTER)
tabel.heading("Jenis", text="Jenis"); tabel.column("Jenis", width=80, anchor=tk.CENTER)
tabel.heading("Uraian", text="Uraian"); tabel.column("Uraian", width=250, anchor=tk.W)
tabel.heading("Nominal", text="Nominal"); tabel.column("Nominal", width=110, anchor=tk.E)
tabel.heading("Rek", text="No. Rek"); tabel.column("Rek", width=130, anchor=tk.W)
tabel.heading("Penerima", text="Penerima"); tabel.column("Penerima", width=130, anchor=tk.W)
tabel.heading("Penyedia", text="Penyedia"); tabel.column("Penyedia", width=130, anchor=tk.W)
tabel.heading("Akun", text="Akun"); tabel.column("Akun", width=150, anchor=tk.CENTER)

scrol = ttk.Scrollbar(frame_tabel, orient="vertical", command=tabel.yview)
tabel.configure(yscroll=scrol.set)
scrol.pack(side=tk.RIGHT, fill=tk.Y)
tabel.pack(fill=tk.BOTH, expand=True)

refresh_tampilan()
root.mainloop()